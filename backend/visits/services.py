import math
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import date, datetime, timedelta
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from django.db.models import Subquery, OuterRef, Q
from django.db.models.functions import Trunc
from visits.models import Session, SessionEntry
from visits.callbacks import statistics_extra_callbacks, StatisticsExtraDataResult


class SessionService:
    """
    Service for managing user sessions and entries.
    It provides methods to enter, exit, update entries, and retrieve session information.
    """

    def enter(self, user: User, type: SessionEntry.SessionEntryType, time: datetime):
        session, _ = Session.objects.get_or_create(user=user, date=timezone.localdate())
        last_entry = session.get_last_entry()

        if not last_entry or not last_entry.end:
            session.add_enter(start=time, type=type)
            return

        # Try to restore session
        break_entry = SessionEntry.objects.create(
            session=session,
            type=SessionEntry.SessionEntryType.BREAK,
            start=last_entry.end,
            end=time,
        )
        break_entry.save()
        session.add_enter(start=time, type=type)

    def update_entry(
        self,
        entry_id: int,
        type: SessionEntry.SessionEntryType,
        start: datetime | None = None,
        end: datetime | None = None,
        comment: str | None = None,
    ):
        entry = get_object_or_404(SessionEntry, id=entry_id)

        for attr, value in {"start": start, "end": end, "comment": comment}.items():
            if value is not None:
                setattr(entry, attr, value)

        entry.type = type
        entry.save()

    def exit(self, user: User, time: datetime, comment: str | None = None):
        session = Session.objects.get_last_user_session(user)
        if session is None:
            raise Session.DoesNotExist()

        last_entry = session.get_last_entry()

        if last_entry is None:
            raise SessionEntry.DoesNotExist("No entry found to exit.")

        if last_entry.end is not None:
            raise ValueError("Entry already checked out.")

        last_entry.end = time
        last_entry.comment = comment
        last_entry.save()

    def handle_leave(
        self,
        user: User,
        type: SessionEntry.SessionEntryType,
        time: datetime,
        comment: str | None = None,
    ):
        session = Session.objects.get_last_user_session(user)
        if session is None:
            raise Session.DoesNotExist()

        last_entry = session.get_last_entry()
        if last_entry is None or last_entry.end is not None:
            raise ValueError("No open work entry to leave from.")

        last_entry.close(time)
        session.add_enter(start=time, type=type, comment=comment)

    def handle_cheater_leave(self, user: User, entry: SessionEntry, end: datetime):
        session: Session = entry.session
        last_entry = session.get_last_entry()

        qs = (
            SessionEntry.objects.filter(session=session)
            .annotate(
                start_trunc=Trunc("start", "minute"), end_trunc=Trunc("end", "minute")
            )
            .filter(Q(end_trunc__gt=entry.start) & Q(start_trunc__lt=end))
            .exclude(pk=entry.pk)
        )

        if last_entry:
            qs = qs.exclude(pk=last_entry.pk)

        if qs.exists():
            raise ValueError("Time overlap with other entries")

        if not last_entry:
            return

        if last_entry.pk == entry.pk:
            last_entry.close(end)
            return

        if last_entry.start < end:
            raise ValueError("Time overlap with other entries")

        entry.close(end)
        SessionEntry.objects.create(
            session=session,
            start=end,
            end=last_entry.start,
            type=SessionEntry.SessionEntryType.BREAK,
        )

    def get_current_session(self, user: User) -> Session | None:
        session = Session.objects.get_last_user_session(user)

        if not session:
            return None

        if session.date == timezone.localdate() or session.get_open_entries().exists():
            return session

        return None

    def get_session_last_comment(self, session: Session | None) -> str | None:
        if session is None:
            return None

        last_entry = session.get_last_entry()
        return last_entry.comment if last_entry else None

    def get_session_status(self, session: Session | None) -> Session.SessionStatus:
        return session.status if session else Session.SessionStatus.INACTIVE

    def get_active_user_with_sessions(self):
        current_sessions = Session.objects.filter(user=OuterRef("pk")).order_by(
            "-date", "-id"
        )

        users = (
            User.objects.filter(is_active=True)
            .annotate(current_session=Subquery(current_sessions.values("id")[:1]))
            .prefetch_related("avatar")
        )

        session_ids = [u.current_session for u in users if u.current_session]
        sessions = Session.objects.filter(id__in=session_ids).select_related("user")
        sessions_by_id = {s.id: s for s in sessions}

        return [
            {"user": user, "session": sessions_by_id.get(user.current_session)}
            for user in users
        ]


class StatisticsService:
    """
    Service for collecting user statistics over a date range.
    It provides methods to retrieve statistics for a user over a specified date range.
    """

    def get_user_date_range_statistics(
        self, user: User, start_date: date, end_date: date
    ) -> list[dict]:
        result = []

        sessions = Session.objects.filter(
            user=user, date__range=(start_date, end_date)
        ).prefetch_related("entries")

        session_date_map = {s.date: s for s in sessions}

        current_date = start_date
        while current_date <= end_date:
            session: Session | None = session_date_map.get(current_date)
            entries = session.entries.all() if session else []
            statistics = self._calculate_statistics(entries)
            extra = self._collect_extra(user, current_date)
            result.append(
                {
                    "date": current_date,
                    "session": session,
                    "statistics": statistics,
                    "extra": extra,
                }
            )
            current_date += timedelta(days=1)

        return result

    def _calculate_statistics(self, entries: list[SessionEntry]) -> dict[str, float]:
        result = {
            "work_time": 0.0,
            "break_time": 0.0,
            "lunch_time": 0.0,
        }

        for entry in entries:
            if not entry.end:
                continue

            delta = (entry.end - entry.start).total_seconds()

            if entry.type == "WORK":
                result["work_time"] += delta
            elif entry.type == "BREAK":
                result["break_time"] += delta
            elif entry.type == "LUNCH":
                result["lunch_time"] += delta

        return result

    def _collect_extra(self, user: User, date: date):
        results: list[StatisticsExtraDataResult] = []
        for callback in statistics_extra_callbacks():
            data = callback(user, date)
            if data:
                results.append(
                    {
                        "type": getattr(callback, "_type"),
                        "payload": data,
                    }
                )

        return results


class XlsxService:
    def user_date_period_statistics_xlsx(
        self, user: User, start: date, end: date, data: list[dict]
    ) -> Workbook:

        def format_timedelta(td: timedelta) -> str:
            total_minutes = int(td.total_seconds() // 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60

            return f"{hours}:{minutes:02d}"

        wb = Workbook()
        ws = wb.active
        if not ws:
            raise Exception

        ws.title = f"Report {user.username} {start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}"

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        gray_fill = PatternFill(
            start_color="CACACA", end_color="CACACA", fill_type="solid"
        )

        ws.append(
            ["Date", "Start Time", "End Time", "Work Time", "Break Time", "Lunch Time"]
        )
        current_idx = 1
        for cell in ws[current_idx]:
            cell.fill = gray_fill

        for row in data:
            entries: list[SessionEntry] = (
                list(row["session"].entries.all()) if row["session"] else []
            )

            if not entries:
                ws.append([row["date"], "--", "--", "--", "--", "--"])
                current_idx += 1
                continue

            first_entry, last_entry = entries[0], entries[-1]
            session_start = (
                first_entry.start.strftime("%H:%M:%S") if first_entry.start else ""
            )
            session_end = last_entry.end.strftime("%H:%M:%S") if last_entry.end else ""

            summary = [
                row["date"],
                session_start,
                session_end,
                format_timedelta(
                    timedelta(seconds=math.floor(row["statistics"]["work_time"] or 0))
                ),
                format_timedelta(
                    timedelta(seconds=math.floor(row["statistics"]["break_time"] or 0))
                ),
                format_timedelta(
                    timedelta(seconds=math.floor(row["statistics"]["lunch_time"] or 0))
                ),
            ]

            ws.append(summary)
            current_idx += 1

            for cell in ws[current_idx]:
                cell.fill = gray_fill

            if len(entries) == 0:
                continue

            ws.append(["", "Start Time", "End Time", "Type", "Comment"])
            current_idx += 1

            ws.row_dimensions[current_idx].outline_level = 1
            ws.row_dimensions[current_idx].hidden = True

            group_start = group_end = current_idx + 1
            for entry in entries:
                ws.append(
                    [
                        "",
                        (entry.start.strftime("%H:%M:%S") if entry.start else ""),
                        entry.end.strftime("%H:%M:%S") if entry.end else "",
                        entry.type if entry.type else "",
                        entry.comment if entry.comment else "",
                    ]
                )
                current_idx += 1
                group_end = current_idx

            for i in range(group_start, group_end + 1):
                ws.row_dimensions[i].outline_level = 1
                ws.row_dimensions[i].hidden = True

        return wb
