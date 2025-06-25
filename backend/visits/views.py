from io import BytesIO
from datetime import date, datetime, timedelta
from math import e
import math
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet
from rest_framework.response import Response
from rest_framework import mixins
from rest_framework.exceptions import APIException, ValidationError, NotFound
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.request import Request
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiResponse
from drf_spectacular.types import OpenApiTypes
from django.utils.translation import gettext as _
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.db.models import Q

from .models import Session, SessionEntry
from .services import SessionService, StatisticsService
from . import serializers


@extend_schema(tags=["visits"])
class EnterView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        "enter",
        request=serializers.SessionEnterSerializer,
        responses={
            status.HTTP_200_OK: None,
            status.HTTP_404_NOT_FOUND: None,
            status.HTTP_500_INTERNAL_SERVER_ERROR: None,
        },
    )
    def post(self, request: Request):
        serializer = serializers.SessionEnterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        start: datetime = serializer.validated_data.get("start")  # type: ignore
        type: SessionEntry.SessionEntryType = serializer.validated_data.get("type")  # type: ignore

        session_service = SessionService()

        try:
            session_service.enter(request.user, type, start)
        except ValueError as e:
            raise ValidationError(detail=str(e))
        except Exception as e:
            raise APIException(detail=str(e))

        return Response(status=status.HTTP_200_OK)


@extend_schema(tags=["visits"])
class ExitView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        "exit",
        request=serializers.SessionExitSerializer,
        responses={
            status.HTTP_200_OK: None,
            status.HTTP_400_BAD_REQUEST: None,
            status.HTTP_404_NOT_FOUND: None,
            status.HTTP_500_INTERNAL_SERVER_ERROR: None,
        },
    )
    def put(self, request: Request):
        serializer = serializers.SessionExitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        time: datetime = serializer.validated_data.get("end")  # type: ignore
        comment: SessionEntry.SessionEntryType = serializer.validated_data.get("comment")  # type: ignore

        session_service = SessionService()

        try:
            session_service.exit(request.user, time, comment)
        except Session.DoesNotExist as e:
            raise NotFound(detail=str(e))
        except ValueError as e:
            raise ValidationError(detail=str(e))
        except Exception as e:
            raise APIException(detail=str(e))

        return Response(status=status.HTTP_200_OK)


@extend_schema(tags=["visits"])
class LeaveView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        "leave",
        request=serializers.SessionEntryLeaveSerializer,
        responses={
            status.HTTP_200_OK: None,
            status.HTTP_404_NOT_FOUND: None,
            status.HTTP_400_BAD_REQUEST: None,
            status.HTTP_500_INTERNAL_SERVER_ERROR: None,
        },
    )
    def post(self, request: Request):
        serializer = serializers.SessionEntryLeaveSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        type: SessionEntry.SessionEntryType = serializer.validated_data.get("type")  # type: ignore
        time: datetime = serializer.validated_data.get("time")  # type: ignore
        comment: str | None = serializer.validated_data.get("comment")  # type: ignore

        session_service = SessionService()

        try:
            session_service.handle_leave(request.user, type, time, comment)
        except Session.DoesNotExist as e:
            raise NotFound(detail=str(e))
        except ValueError as e:
            raise ValidationError(detail=str(e))
        except Exception as e:
            raise APIException(detail=str(e))

        return Response()


@extend_schema(tags=["visits"])
@extend_schema_view(
    create=extend_schema("createEntry"),
    destroy=extend_schema("destroyEntry"),
    partial_update=extend_schema("partialUpdateEntry"),
    update=extend_schema("updateEntry"),
)
class SessionEntryModelViewset(
    GenericViewSet,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
):
    permission_classes = [IsAuthenticated]
    serializer_class = serializers.SessionEntryModelSerializer

    def get_queryset(self):
        return SessionEntry.objects.filter(session__user=self.request.user)

    def perform_create(self, serializer):
        session_id = self.kwargs.get("session_id")
        if not session_id:
            raise ValidationError("Session id was not provided")

        session = get_object_or_404(Session, id=session_id, user=self.request.user)

        new_start = serializer.validated_data.get("start")
        new_end = serializer.validated_data.get("end")

        conflict_exists = (
            SessionEntry.objects.filter(session=session)
            .filter(Q(start__lt=new_end) & Q(end__gt=new_start))
            .exists()
        )

        if conflict_exists:
            raise ValidationError("New entry overlaps with an existing entry.")

        serializer.save(session=session)


@extend_schema(tags=["visits"])
class CurrentSessionView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        "current",
        summary="Get current session info",
        responses={
            status.HTTP_200_OK: serializers.SessionModelSerializer,
            status.HTTP_404_NOT_FOUND: None,
        },
    )
    def get(self, request: Request):
        session_service = SessionService()
        session = session_service.get_current_session(request.user)
        if session is None:
            return Response({"status": Session.SessionStatus.INACTIVE, "entries": []})

        serializer = serializers.SessionModelSerializer(session)

        return Response(serializer.data)


@extend_schema(tags=["visits"])
class UsersTodayView(APIView):
    permission_classes = [AllowAny]

    @extend_schema(
        "today",
        responses={status.HTTP_200_OK: serializers.UserSessionSerializer(many=True)},
    )
    def get(self, request: Request):

        session_service = SessionService()
        active_users_with_sessions = session_service.get_active_user_with_sessions()

        result = []
        for user_session in active_users_with_sessions:
            user: User = user_session.get("user")  # type: ignore
            session: Session | None = user_session.get("session")

            result.append(
                {
                    "user": user,
                    "session": {
                        "status": session_service.get_session_status(session),
                        "comment": session_service.get_session_last_comment(session),
                    },
                }
            )

        serializer = serializers.UserSessionSerializer(
            result, many=True, context={"request": request}
        )
        return Response(serializer.data)


@extend_schema(tags=["statistics"])
class UserMonthStatisticsView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        "statistics",
        parameters=[serializers.UserMonthStatisticsRequestSerializer],
        responses=serializers.UserMonthStatisticsResponseSerializer(many=True),
    )
    def get(self, request: Request):
        request_serializer = serializers.UserMonthStatisticsRequestSerializer(
            data=request.query_params
        )
        request_serializer.is_valid(raise_exception=True)
        start: date = request_serializer.validated_data["start"]  # type: ignore
        end: date = request_serializer.validated_data["end"]  # type: ignore
        user = request.user

        statistics_service = StatisticsService()
        result = statistics_service.get_user_date_range_statistics(user, start, end)

        response_serializer = serializers.UserMonthStatisticsResponseSerializer(
            result, many=True
        )

        return Response(response_serializer.data)


@extend_schema(tags=["statistics"])
class ExportUserReportView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        "export",
        parameters=[serializers.UserMonthStatisticsRequestSerializer],
        responses={
            status.HTTP_200_OK: OpenApiResponse(
                response=OpenApiTypes.BINARY,
                description="Excel file with user statistics",
            )
        },
    )
    def get(self, request: Request):
        request_serializer = serializers.UserMonthStatisticsRequestSerializer(
            data=request.query_params
        )
        request_serializer.is_valid(raise_exception=True)
        start: date = request_serializer.validated_data["start"]  # type: ignore
        end: date = request_serializer.validated_data["end"]  # type: ignore
        user = request.user

        statistics_service = StatisticsService()
        result = statistics_service.get_user_date_range_statistics(user, start, end)
        wb = self._prepare_report_sheet(user, start, end, result)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{request.user.username} {str(start.strftime("%Y-%m-%d"))} {str(end.strftime("%Y-%m-%d"))}.xlsx"'
            },
        )

        return response

    def _prepare_report_sheet(
        self, user: User, start: date, end: date, data: list[dict]
    ) -> Workbook:

        def format_timedelta(td: timedelta) -> str:
            total_minutes = int(td.total_seconds() // 60)
            hours = total_minutes // 60
            minutes = total_minutes % 60

            return f"{hours}:{minutes:02d}"

        wb = Workbook()
        ws = wb.active
        if not ws:
            raise APIException(_("Failed to create worksheet."))

        ws.title = f"Report {user.username} {start.strftime('%Y-%m-%d')} - {end.strftime('%Y-%m-%d')}"

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        gray_fill = PatternFill(
            start_color="CACACA", end_color="CACACA", fill_type="solid"
        )

        ws.append(
            ["Date", "Start Time", "End Time", "Work Time", "Break Time", "Lunch Time"]
        )
        current_idx = 1
        for cell in ws[current_idx]:
            cell.fill = gray_fill

        for row in data:
            entries: list[SessionEntry] = (
                list(row["session"].entries.all()) if row["session"] else []
            )

            if not entries:
                ws.append([row["date"], "--", "--", "--", "--", "--"])
                current_idx += 1
                continue

            first_entry, last_entry = entries[0], entries[-1]
            session_start = (
                first_entry.start.strftime("%H:%M:%S") if first_entry.start else ""
            )
            session_end = last_entry.end.strftime("%H:%M:%S") if last_entry.end else ""

            summary = [
                row["date"],
                session_start,
                session_end,
                format_timedelta(
                    timedelta(seconds=math.floor(row["statistics"]["work_time"] or 0))
                ),
                format_timedelta(
                    timedelta(seconds=math.floor(row["statistics"]["break_time"] or 0))
                ),
                format_timedelta(
                    timedelta(seconds=math.floor(row["statistics"]["lunch_time"] or 0))
                ),
            ]

            ws.append(summary)
            current_idx += 1

            for cell in ws[current_idx]:
                cell.fill = gray_fill

            if len(entries) == 0:
                continue

            ws.append(["", "Start Time", "End Time", "Type", "Comment"])
            current_idx += 1

            ws.row_dimensions[current_idx].outline_level = 1
            ws.row_dimensions[current_idx].hidden = True

            group_start = group_end = current_idx + 1
            for entry in entries:
                ws.append(
                    [
                        "",
                        (entry.start.strftime("%H:%M:%S") if entry.start else ""),
                        entry.end.strftime("%H:%M:%S") if entry.end else "",
                        entry.type if entry.type else "",
                        entry.comment if entry.comment else "",
                    ]
                )
                current_idx += 1
                group_end = current_idx

            for i in range(group_start, group_end + 1):
                ws.row_dimensions[i].outline_level = 1
                ws.row_dimensions[i].hidden = True

        return wb
